"""Append job application records to an Excel file across runs."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from src.utils.formatters import excel_cell_value

SHEET_NAME = "Naukri Jobs"

COLUMNS = [
    "Run Date",
    "Run Time",
    "Job ID",
    "Job Title",
    "Company",
    "Location",
    "Experience",
    "Salary",
    "Posted Date",
    "Search Keyword",
    "Skills",
    "Job URL",
    "Status",
    "Applied At",
    "Notes",
]


class ExcelJobLogger:
    """Creates or appends rows to a single Excel workbook."""

    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        self.run_date = datetime.now().strftime("%Y-%m-%d")
        self.run_time = datetime.now().strftime("%H:%M:%S")

    def load_applied_job_ids(self) -> set[str]:
        """Return job IDs that were successfully applied in any previous run."""
        if not self.filepath.exists():
            return set()

        wb = load_workbook(self.filepath, read_only=True, data_only=True)
        try:
            ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            job_id_idx = headers.index("Job ID")
            status_idx = headers.index("Status")

            applied: set[str] = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= max(job_id_idx, status_idx):
                    continue
                if row[status_idx] == "Applied" and row[job_id_idx]:
                    applied.add(str(row[job_id_idx]))
            return applied
        finally:
            wb.close()

    def append_job(
        self,
        job,
        keyword: str,
        status: str,
        notes: str = "",
        applied_at: str = "",
    ) -> None:
        wb = self._open_or_create_workbook()
        ws = wb[SHEET_NAME]

        skills = ", ".join(job.tags) if getattr(job, "tags", None) else ""
        apply_link = getattr(job, "apply_link", "") or f"https://www.naukri.com/job-listings-{job.job_id}"

        ws.append(
            [
                excel_cell_value(self.run_date),
                excel_cell_value(self.run_time),
                excel_cell_value(job.job_id),
                excel_cell_value(job.title),
                excel_cell_value(job.company),
                excel_cell_value(job.location),
                excel_cell_value(job.experience),
                excel_cell_value(job.salary),
                excel_cell_value(job.posted_date),
                excel_cell_value(keyword),
                excel_cell_value(skills),
                excel_cell_value(apply_link),
                excel_cell_value(status),
                excel_cell_value(applied_at),
                excel_cell_value(notes),
            ]
        )

        wb.save(self.filepath)

    def append_run_summary(self, stats: dict, total_found: int) -> None:
        wb = self._open_or_create_workbook()
        ws = wb[SHEET_NAME]

        ws.append(
            [
                self.run_date,
                self.run_time,
                "",
                "=== RUN SUMMARY ===",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "Summary",
                "",
                (
                    f"Found={total_found}; Applied={stats.get('applied', 0)}; "
                    f"Already Applied={stats.get('skipped_applied', 0)}; "
                    f"External={stats.get('skipped_external', 0)}; "
                    f"Failed={stats.get('failed', 0)}"
                ),
            ]
        )
        wb.save(self.filepath)

    def _open_or_create_workbook(self):
        if self.filepath.exists():
            return load_workbook(self.filepath)

        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(COLUMNS)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        wb.save(self.filepath)
        return wb
